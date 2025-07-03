import openpyxl
from django.http import HttpResponse
from .models import Pago, Reserva
from django.db.models import Sum
from collections import Counter

def obtener_datos_reporte_ventas():
    # Total ingresos
    total_ingresos = Pago.objects.filter(estado='completado').aggregate(total=Sum('monto'))['total'] or 0

    # Ingresos por sala
    ingresos_por_sala = Pago.objects.filter(estado='completado') \
        .values('reserva__sala__nombre') \
        .annotate(total=Sum('monto')) \
        .order_by('-total')

    # Total reservas
    total_reservas = Reserva.objects.filter(pago__estado='completado').count()

    # Conteo de reservas por película (manual)
    reservas = Reserva.objects.filter(pago__estado='completado').select_related('sala')
    contador_peliculas = Counter()

    for reserva in reservas:
        peliculas = reserva.sala.peliculas.all()
        for pelicula in peliculas:
            contador_peliculas[pelicula.titulo] += 1

    reservas_por_pelicula = [{'titulo': titulo, 'cantidad': cantidad} for titulo, cantidad in contador_peliculas.items()]

    return {
        'total_ingresos': total_ingresos,
        'ingresos_por_sala': ingresos_por_sala,
        'total_reservas': total_reservas,
        'reservas_por_pelicula': reservas_por_pelicula,
    }

def generar_excel_reporte_ventas():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from io import BytesIO
    import requests
    from django.http import HttpResponse
    import datetime

    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Configuración página
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    
    # Logo
    try:
        response = requests.get("https://cdn-icons-png.flaticon.com/512/2455/2455201.png")
        img = Image(BytesIO(response.content))
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
    except:
        pass
    
    # Título
    ws.merge_cells('C1:H3')
    title_cell = ws['C1']
    title_cell.value = "REPORTE DE VENTAS - CINEBOX"
    title_cell.font = Font(name='Calibri', size=24, bold=True, color='1F4E78')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fechas
    ws['C4'] = "Fecha de generación:"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    ws['C5'] = "Periodo reportado:"
    ws['C5'].font = Font(bold=True)
    ws['D5'] = "01/01/2025 - 30/06/2025"
    
    # Encabezados
    headers = ['ID Pago', 'Usuario', 'Sala', 'Monto', 'Método', 'Estado', 'Fecha de Pago']
    ws.append([''])  # Espacio
    ws.append(headers)
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    pagos = Pago.objects.filter(estado='completado').select_related('reserva', 'reserva__usuario', 'reserva__sala')
    
    even_row_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    
    row_num = 8
    total_ventas = 0
    
    for pago in pagos:
        usuario = pago.reserva.usuario.email if pago.reserva and pago.reserva.usuario else ''
        sala = pago.reserva.sala.nombre if pago.reserva and pago.reserva.sala else ''
        monto = float(pago.monto)
        fecha_pago = pago.fecha_pago.strftime('%d/%m/%Y %H:%M') if pago.fecha_pago else ''
        
        ws.append([
            pago.id,
            usuario,
            sala,
            monto,
            pago.metodo_pago,
            pago.estado.capitalize(),
            fecha_pago
        ])
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
        
        ws.cell(row=row_num, column=4).number_format = '"S/"#,##0.00'
        ws.cell(row=row_num, column=7).number_format = 'dd/mm/yyyy hh:mm'
        
        total_ventas += monto
        row_num += 1
    
    # Totales
    ws.append([''] * len(headers))
    total_row = row_num
    
    ws.cell(row=total_row, column=3, value="TOTAL VENTAS:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_ventas).number_format = '"S/"#,##0.00'
    ws.cell(row=total_row, column=4).font = Font(bold=True, color='1F4E78')
    
    ws.cell(row=total_row+1, column=3, value="Total de transacciones:").font = Font(bold=True)
    ws.cell(row=total_row+1, column=4, value=len(pagos)).font = Font(bold=True)
    
    ws.cell(row=total_row+2, column=3, value="Ticket promedio:").font = Font(bold=True)
    ws.cell(row=total_row+2, column=4, value=total_ventas/len(pagos) if pagos else 0).number_format = '"S/"#,##0.00'
    ws.cell(row=total_row+2, column=4).font = Font(bold=True)
    
    # Autoajuste columnas
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Config columnas específicas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    
    # Gráfico
    try:
        from collections import defaultdict
        from openpyxl.chart import BarChart, Reference

        ventas_por_sala = defaultdict(float)
        for pago in pagos:
            sala_nombre = pago.reserva.sala.nombre if pago.reserva and pago.reserva.sala else 'Sin Sala'
            ventas_por_sala[sala_nombre] += float(pago.monto)

        chart_sheet = wb.create_sheet(title="Resumen por Sala")
        chart_sheet.append(['Sala', 'Ventas'])
        for sala, ventas in ventas_por_sala.items():
            chart_sheet.append([sala, ventas])

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Ventas por Sala"
        chart.y_axis.title = 'Soles (S/)'
        chart.x_axis.title = 'Sala'

        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(ventas_por_sala)+1)
        categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(ventas_por_sala)+1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, "I5")
    except Exception as e:
        print(f"Error al crear gráfico: {e}")
    
    # Congelar encabezados
    ws.freeze_panes = 'A8'
    
    # Pie de página solo con texto
    ws.oddFooter.center.text = "CineMax - Reporte de Ventas"
    ws.oddFooter.right.text = "Página &P de &N"
    
    # Descargar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas_CineMax.xlsx'
    wb.save(response)
    
    return response
